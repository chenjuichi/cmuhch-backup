import os
import datetime

import openpyxl

from openpyxl import Workbook, load_workbook
# 載入 Font 和 PatternFill 模組
from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------

from flask import Blueprint, jsonify, request
from sqlalchemy import func

excelTable = Blueprint('excelTable', __name__)

# ------------------------------------------------------------------


@excelTable.route("/exportToExcelForReq", methods=['POST'])
def export_to_Excel_for_Req():
    print("exportToExcelForReq....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    _name = request_data['name']
    temp = len(_blocks)
    data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    date = datetime.datetime.now()
    today = date.strftime('%Y-%m-%d-%H%M')
    current_file = '..\\領用記錄查詢_'+today + '.xlsx'
    print("filename:", current_file)
    file_check = os.path.exists(current_file)  # true:excel file exist

    return_value = True  # true: export into excel成功
    if not data_check:  # false: 資料不完全
        return_value = False

    if return_value:
        if file_check:
            wb = openpyxl.load_workbook(current_file)  # 開啟現有的 Excel 活頁簿物件
        else:
            wb = Workbook()     # 建立空白的 Excel 活頁簿物件
        # ws = wb.active
        ws = wb.worksheets[0]   # 取得第一個工作表

        ws.title = '領用記錄查詢-' + _name                # 修改工作表 1 的名稱為 oxxo
        ws.sheet_properties.tabColor = '7da797'  # 修改工作表 1 頁籤顏色為紅色

        for obj in _blocks:
            temp_array = []
            temp_array.append(obj['reqRecord_reagID'])
            temp_array.append(obj['reqRecord_reagName'])
            temp_array.append(obj['reqRecord_supplier'])
            temp_array.append(obj['reqRecord_stockInDate'])
            temp_array.append(obj['reqRecord_Date'])
            temp_array.append(obj['reqRecord_Employer'])
            # temp_str = obj['reqRecord_cnt'] + obj['reqRecord_unit']
            temp_array.append(obj['reqRecord_cnt'])

            ws.append(temp_array)
            #print("obj: ", temp_array)

        for col in ws.columns:
            column = col[0].column_letter  # Get the column name
            temp_cell = column + '1'
            ws[temp_cell].font = Font(
                name='微軟正黑體', color='ff0000', bold=True)  # 設定儲存格的文字樣式
            ws[temp_cell].alignment = Alignment(horizontal='center')
            ws.column_dimensions[column].bestFit = True
            # dim = openpyxl.worksheet.dimensions.ColumnDimension(
            #    ws, index=column, bestFit=True, customWidth=True)
            # ws.column_dimensions[column] = dim

        wb.save(current_file)

    return jsonify({
        'status': return_value,
        'outputs': current_file,
    })


@excelTable.route("/exportToExcelForStock", methods=['POST'])
def export_to_Excel_for_Stock():
    print("exportToExcelForStock....")

    request_data = request.get_json()

    _blocks = request_data['blocks']
    _count = request_data['count']
    _name = request_data['name']
    temp = len(_blocks)
    data_check = (True, False)[_count == 0 or temp == 0 or _count != temp]

    date = datetime.datetime.now()
    today = date.strftime('%Y-%m-%d-%H%M')
    current_file = '..\\庫存記錄查詢_'+today + '.xlsx'
    print("filename:", current_file)
    file_check = os.path.exists(current_file)  # true:excel file exist

    return_value = True  # true: export into excel成功
    if not data_check:  # false: 資料不完全
        return_value = False

    if return_value:
        if file_check:
            wb = openpyxl.load_workbook(current_file)  # 開啟現有的 Excel 活頁簿物件
        else:
            wb = Workbook()     # 建立空白的 Excel 活頁簿物件
        # ws = wb.active
        ws = wb.worksheets[0]   # 取得第一個工作表

        ws.title = '庫存記錄查詢-' + _name                # 修改工作表 1 的名稱為 oxxo
        ws.sheet_properties.tabColor = '7da797'  # 修改工作表 1 頁籤顏色為紅色

        for obj in _blocks:
            temp_array = []

            temp_array.append(obj['stkRecord_reagID'])
            temp_array.append(obj['stkRecord_reagName'])
            temp_array.append(obj['stkRecord_supplier'])
            temp_array.append(obj['stkRecord_Date'])  # 入庫日期
            temp_array.append(obj['stkRecord_period'])  # 效期
            temp_array.append(obj['stkRecord_saftStockUnit'])
            temp_array.append(obj['stkRecord_cntUnit'])
            temp_array.append(obj['stkRecord_grid'])

            ws.append(temp_array)
            #print("obj: ", temp_array)

        for col in ws.columns:
            column = col[0].column_letter  # Get the column name
            temp_cell = column + '1'
            ws[temp_cell].font = Font(
                name='微軟正黑體', color='ff0000', bold=True)  # 設定儲存格的文字樣式
            ws[temp_cell].alignment = Alignment(horizontal='center')
            ws.column_dimensions[column].bestFit = True

        wb.save(current_file)

    return jsonify({
        'status': return_value,
        'outputs': current_file,
    })
